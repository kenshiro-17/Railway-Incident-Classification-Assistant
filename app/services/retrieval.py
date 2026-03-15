from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook


@dataclass
class HistoricalIncident:
    incident_id: str
    class_label: str
    symptoms: str
    resolution: str


class HistoricalRetrievalStore:
    def __init__(self) -> None:
        self.rows: List[HistoricalIncident] = []
        self.incident_ids: set[str] = set()

    def ingest_csv(self, path: str) -> Dict[str, int]:
        added = 0
        rejected = 0
        duplicate = 0
        with Path(path).open("r", encoding="utf-8") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                if not self._valid(row):
                    rejected += 1
                    continue
                incident_id = str(row["incident_id"])
                if incident_id in self.incident_ids:
                    rejected += 1
                    duplicate += 1
                    continue
                self.rows.append(
                    HistoricalIncident(
                        incident_id=incident_id,
                        class_label=row["class_label"],
                        symptoms=row["symptoms"],
                        resolution=row.get("resolution", "")
                    )
                )
                self.incident_ids.add(incident_id)
                added += 1
        return {"added": added, "rejected": rejected, "duplicate": duplicate}

    def ingest_xlsx(self, path: str) -> Dict[str, int]:
        wb = load_workbook(path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = [str(h) for h in rows[0]]
        idx = {name: i for i, name in enumerate(header)}
        added = 0
        rejected = 0
        duplicate = 0
        for row in rows[1:]:
            mapped = {k: row[i] for k, i in idx.items()}
            if not self._valid(mapped):
                rejected += 1
                continue
            incident_id = str(mapped["incident_id"])
            if incident_id in self.incident_ids:
                rejected += 1
                duplicate += 1
                continue
            self.rows.append(
                HistoricalIncident(
                    incident_id=incident_id,
                    class_label=str(mapped["class_label"]),
                    symptoms=str(mapped["symptoms"]),
                    resolution=str(mapped.get("resolution", ""))
                )
            )
            self.incident_ids.add(incident_id)
            added += 1
        return {"added": added, "rejected": rejected, "duplicate": duplicate}

    def ingest_eval_dataset_json(self, path: str) -> Dict[str, int]:
        added = 0
        rejected = 0
        duplicate = 0
        payload = json.loads(Path(path).read_text(encoding="utf-8"))
        scenarios = payload.get("scenarios", [])
        for scenario in scenarios:
            scenario_id = str(scenario.get("scenario_id", "")).strip()
            class_label = str(scenario.get("expected_class", "")).strip()
            symptoms = str(scenario.get("symptoms", "")).strip()
            if not scenario_id or not class_label or not symptoms:
                rejected += 1
                continue
            if scenario_id in self.incident_ids:
                duplicate += 1
                rejected += 1
                continue
            self.rows.append(
                HistoricalIncident(
                    incident_id=scenario_id,
                    class_label=class_label,
                    symptoms=symptoms,
                    resolution="",
                )
            )
            self.incident_ids.add(scenario_id)
            added += 1
        return {"added": added, "rejected": rejected, "duplicate": duplicate}

    def _valid(self, row: dict) -> bool:
        return bool(row.get("incident_id") and row.get("class_label") and row.get("symptoms"))

    def similar(self, symptoms: str, top_k: int = 3) -> List[tuple[HistoricalIncident, float]]:
        query_tokens = set(symptoms.lower().split())

        def score(item: HistoricalIncident) -> float:
            tokens = set(item.symptoms.lower().split())
            if not tokens:
                return 0.0
            return len(tokens & query_tokens) / max(1, len(tokens | query_tokens))

        ranked = sorted(((item, score(item)) for item in self.rows), key=lambda x: x[1], reverse=True)
        return ranked[:top_k]


retrieval_store = HistoricalRetrievalStore()
